#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Modified https://xdevs.com/doc/Wavetek/4950/calkit_4950.py for python3
# and for use with a GPIB adapter than can only have 1 open resource at a time.
# Needs formatted https://xdevs.com/doc/Wavetek/4950/test_4950.xlsx to be present

import sys
import pyvisa as visa
import time
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.styles import colors, Font, Fill, NamedStyle

rm = visa.ResourceManager()
#F5700EP = rm.open_resource('GPIB0::1::INSTR') # Ethernet GPIB Dongle
#dmm = rm.open_resource('GPIB0::9::INSTR') # Ethernet GPIB Dongle

F5700EP = rm.open_resource('GPIB0::1::INSTR') # Local GPIB Dongle
#dmm = rm.open_resource('GPIB0::9::INSTR') # Ethernet GPIB Dongle

########## DMM and MFC ##########   
F5700EP.write("*RST")
F5700EP.write("*CLS")
F5700EP.write("RANGELCK OFF")
#F5700EP.write("OUT 10 V, 0 Hz")
#F5700EP.write("RANGELCK ON")
F5700EP.write("STBY") 
F5700EP.write("OUT 0.0 V, 0 Hz")
F5700EP.write("EXTGUARD OFF")
time.sleep(5)
F5700EP.write("OPER")
time.sleep(5)
print("SRC configured")

F5700EP.close()
dmm = rm.open_resource('GPIB0::9::INSTR')
dmm.timeout = None
dmm.write("*RST")
time.sleep(2)
dmm.write("DCV 1000,PCent_100")
dmm.write("TRIG_SRCE EXT")    
dmm.write("ACCURACY HIGH")
dmm.write("CORRECTN CERTIFIED")
dmm.write("BAND OFF")
dmm.write("LCL_GUARD")
dmm.write("LEAD_NO '31255'")
dmm.timeout = 100000
print("D4950 configured")
   
wb = load_workbook('test_4950.xlsx')
ws = wb.active
time_start = time.time()

########## Colecting info of DMM & Calibrator ##########
########## info of DMM ##########

cutstr = []   
dmm.write("*IDN?")
info = dmm.read()
cutstr = info.split(",")
ws['C5'] = cutstr[0] #Manufacturer
ws['C6'] = cutstr[1]  #Model Number
ws['C7'] = 'Datron 4950 STD' #Meter info######
dmm.write("*OPT?")
info = dmm.read()
ws['C8'] = info #Options
ws['C9'] = cutstr[3] #Firmware
dmm.write("DATE? BASE")
info = dmm.read()
ws['C10'] = info # date of baseline calibration
dmm.write("DATE? CERTIFIED")
info = dmm.read()
ws['C11'] = info # date of CERTIFIED calibration
#dmm.write("TST?")
#time.sleep(600)
#info = dmm.read()
#ws['C12'] = info # Fast Test: returns '0' if OK,otherwise '1' 


ws['H5'] = cutstr[2] #S/N######
ws['H6'] = time.strftime("%y-%m-%d %H:%M:%S")#Test date
ws['H7'] = '23.0' #Ambient Temperature N######
dmm.write("CERT_AMB?")
info = float(dmm.read())
ws['H8'] = info #CERTIFIED Ambient temp.
time.sleep(1)
dmm.write("TEMP?")
info = float(dmm.read())
ws['H9'] = info #
time.sleep(1)
dmm.write("TEMP? BASE")
info = float(dmm.read())
ws['H10'] = info #
time.sleep(1)
dmm.write("TEMP? CERTIFIED")
info = float(dmm.read())
ws['H11'] = info #
dmm.write("STANDBY?")
info = dmm.read()
ws['H12'] = info #

########## info of Calibrator ##########
ws['A21'] = 'Fluke 5700A' #MFC
ws['C21'] = '' #Option
ws['E21'] = '--' #Unc
ws['F21'] = '--' #Calibraton Date
ws['H21'] = '--' #Due Date
ws['A22'] = 'ADRmu12' #Voltage reference
ws['C22'] = '10.000000 VDC' #Option
ws['E22'] = '--' #Unc
ws['F22'] = '--' #Calibration Date
ws['H22'] = '--' #Due Date
ws['A23'] = 'Fluke 742A-1 8927002' #Resistor standard
ws['C23'] = '1.0000195 ohm' #Option
ws['E23'] = '--' #Unc
ws['F23'] = '--' #Calibration Date
ws['H23'] = '--' #Due Date

ws['A24'] = 'Fluke 742A-10k 7815006' #Resistor standard
ws['C24'] = '10.000133 kohm'#Option
ws['E24'] = '' #Unc
ws['F24'] = '' #Calibration Date
ws['H24'] = '' #Due Date

dmm.close()
F5700EP = rm.open_resource('GPIB0::1::INSTR')

########## info of MFC ##########
F5700EP.write("CAL_DATE? CAL")
info = F5700EP.read()
ws['C29'] = info #
F5700EP.write("CAL_DATE? WBFLAT")
info = F5700EP.read()
ws['C30'] = info #
F5700EP.write("CAL_DATE? WBGAIN")
info = F5700EP.read()
ws['C31'] = info #
F5700EP.write("CAL_DATE? ZERO")
info = F5700EP.read()
ws['C32'] = info #
#F5700EP.write("CAL_CONF?")#NOT Suport in 5700 
#info = F5700EP.read()
ws['C33'] = '5700A 99% 90 Days' #MFC Confidence level N######
F5700EP.write("CAL_CONST? CHECK, KV6")
info = F5700EP.read()
ws['C34'] = info #CAL CONST 6.5V reference voltage
F5700EP.write("CAL_CONST? CHECK, KV13")
info = F5700EP.read()
ws['C35'] = info #CAL CONST 13V reference voltage
F5700EP.write("CAL_CONST? CHECK, R10K")
info = F5700EP.read()
ws['C36'] = info #CAL CONST 10KOHM true output resistance
F5700EP.write("CAL_CONST? CHECK, RS10K")
info = F5700EP.read()
ws['C37'] = info #CAL CONST 10KOHM standard resistance
F5700EP.write("CAL_CONST? CHECK, R1")
info = F5700EP.read()
ws['C38'] = info #CAL CONST 1 OHM true output resistance
#total number of minutes that the power switch has been on
F5700EP.write("ETIME?")
info = float(F5700EP.read())
ws['C39'] = ("%d days " % (info//1440)) + ("%.2f hours" % ((info%1440)/60))
#Returns the number of minutes since power-up this session 
F5700EP.write("ONTIME?")
info = float(F5700EP.read())
ws['C40'] = ("%d days " % (info//1440)) + ("%.2f hours" % ((info%1440)/60))

F5700EP.write("CAL_DAYS? CAL")
info = F5700EP.read()
ws['H29'] = ("%s days" % (info)) #
F5700EP.write("CAL_DAYS? WBFLAT")
info = F5700EP.read()
ws['H30'] = ("%s days" % (info)) #
F5700EP.write("CAL_DAYS? WBGAIN")
info = F5700EP.read()
ws['H31'] = ("%s days" % (info)) #
F5700EP.write("CAL_DAYS? ZERO")
info = F5700EP.read()
ws['H32'] = ("%s days" % (info)) #
F5700EP.write("CAL_CONST? CHECK, D3P")
info = float(F5700EP.read())
ws['H33'] = float("%.7f" % (info)) # #11V range, positive zero
F5700EP.write("CAL_CONST? CHECK, D3M")
info = float(F5700EP.read())
ws['H34'] = float("%.7f" % (info))  #11V range, negative zero
F5700EP.write("CAL_CONST? CHECK, ZERO_TEMP")
info = float(F5700EP.read())
ws['H35'] = float("%.3f" % (info))  #
F5700EP.write("CAL_CONST? CHECK, ALL_TEMP")
info = float(F5700EP.read())
ws['H36'] = float("%.3f" % (info))  #
F5700EP.write("CAL_CONST? CHECK, WB_FLAT_TEMP")
info = float(F5700EP.read())
ws['H37'] = float("%.3f" % (info))  #

F5700EP.write("CAL_CONST? CHECK, DACLIN")
info = float(F5700EP.read())
ws['H40'] = float("%.3f" % (info)) #

#to be continue


ws['B373'] = float("%.2f" % ((time.time() - time_start)/60))
wb.save('test_4950.xlsx')
print("Initialization time elapsed:%.2f minutes"
      % ((time.time() - time_start)/60))

time_start = time.time()

F5700EP.close()


########## DCV PERFORMANCE TEST ##########
### SHORT Input ###
dmm = rm.open_resource('GPIB0::9::INSTR')
dmm.timeout = None
dmm.write("*TRG;GET;RDG?")
volt = float(dmm.read())
if abs(volt) > 1e-3:
    print ("Smoke elapsedd, check connections")
    quit()

#DCV test Zero input
print("OFFSET TESTS (NOTE: Math Null is Disabled)")
dmm.write("DCV 0.1,PCENT_0,LCL_GUARD")
dmm.write("ZERO?")
time.sleep(100)

for ix in range (0,5):
    array = []
    sdev = 0.0
    median = 0.0
    unc = 0
    Quality = 0
    cutstr = []
    if ix == 0:
        dmm.write("DCV 0.1,PCENT_0,LCL_GUARD")
    if ix == 1:
        dmm.write("DCV 1,PCENT_0,LCL_GUARD")
    if ix == 2:
        dmm.write("DCV 10,PCENT_0,LCL_GUARD")
    if ix == 3:
        dmm.write("DCV 100,PCENT_0,LCL_GUARD")
    if ix == 4:
        dmm.write("DCV 1000,PCENT_0,LCL_GUARD")
    for i in range (0,10) :
        dmm.write("*TRG;GET;RDG?")
        volt = float(dmm.read())
        array.extend([volt])
    sdev = np.std(array[1:],ddof = 1)
    median = np.median(array[1:])
    ws['C' + str(45+ix)] = median*1e6
    ws['J' + str(45+ix)] = sdev*1e6
    
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    
    F5700EP.write("UNCERT?")
    unc = F5700EP.read()
    cutstr = unc.split(",")
    ws['D' + str(45+ix)] = float(cutstr[0])
    
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None    
    dmm.write("DEVTN? ABSOLUTE")
    Quality = float(dmm.read())
    ws['K' + str(45+ix)] = Quality*1e6        
    print("Source = %s, dmm = %.10f V,  sdev = %.3f uV, Quality(DMM) = %.3f uV" % ("0 V", median, sdev*1e6, Quality*1e6))
    print(array)
wb.save('test_4950.xlsx')

### DCV test 0.01V to 1000V ###
print("DCV GAIN TEST")
DCV_list = [0.01,0.02,0.05,0.10,-0.01,-0.02,-0.05,-0.10,
            0.10,0.20,0.50,1.00,-0.10,-0.20,-0.50,-1.00,
            1.00,2.00,5.00,10.0,19.0,-1.00,-2.00,-5.00,-10.0,-19.0,
            10.0,20.0,50.0,100.0,-10.0,-20.0,-50.0,-100.0,
            -100.0,-200.0,-500.0,-1000.0,100.0,200.0,500.0,1000.0]

dmm.write("DCV 1000,PCENT_100,LCL_GUARD")    
dmm.close()
for ix in range (0,42):
    
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None    
    array = []
    sdev = 0.0
    median = 0.0
    unc = 0
    Quality = 0
    vout = float(DCV_list[ix])
    if abs(vout) <= 0.1 and ix < 8:
        dmm.write("DCV 0.1,PCENT_100,LCL_GUARD")
        print("DMM DCV Range: 100mV")
    elif (abs(vout) >= 0.1 and abs(vout) <= 1) and (ix >= 8 and ix < 16):
        dmm.write("DCV 1,PCENT_100,LCL_GUARD")
        print("DMM DCV Range: 1V")
    elif (abs(vout) >= 1 and abs(vout) <= 19) and (ix >= 16 and ix < 26):
        dmm.write("DCV 10,PCENT_100,LCL_GUARD")
        print("DMM DCV Range: 10V")
    elif (abs(vout) >= 10 and abs(vout) <= 100) and (ix >= 26 and ix < 34):
        dmm.write("DCV 100,PCENT_100,LCL_GUARD")
        print("DMM DCV Range: 100V")
    elif (abs(vout) >= 100 and abs(vout) <= 1000) and (ix >= 34 and ix < 42):
        dmm.write("DCV 1000,PCENT_100,LCL_GUARD")
        print("DMM DCV Range: 1000V")
    else:
        dmm.write("DCV 1000,PCENT_100,LCL_GUARD")
        
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
        
    F5700EP.write("OUT %.7f" % vout)
    F5700EP.write("OPER")
    
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None    
    if ix == 36 or ix == 37 or ix >= 40:
        dmm.write("TRIG_SRCE INT")
        time.sleep(500) # waiting 300s at 500 and 1000V
        dmm.write("TRIG_SRCE EXT")        
    else:
        time.sleep(10)
    for i in range (0,10):
        dmm.write("*TRG;GET;RDG?")
        volt = float(dmm.read())
        array.extend([volt])
    sdev = np.std(array[1:],ddof = 1)
    median = np.median(array[1:])
    ws['C' + str(51+ix)] = median
    ws['J' + str(51+ix)] = sdev*1e6/median
    
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    
    F5700EP.write("UNCERT?")
    unc = F5700EP.read()
    cutstr = unc.split(",")
    if str(cutstr[1]) == "V":
        ws['D' + str(51+ix)] = float(cutstr[0])*1e6/vout
    else:
        ws['D' + str(51+ix)] = float(cutstr[0])
        
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None        
    dmm.write("DEVTN? READING")
    Quality = float(dmm.read())
    ws['K' + str(51+ix)] = Quality    
    print("Source = %.3f V, dmm = %.8f V,  sdev = %.3f ppm, Quality(DMM) =  %s" % (vout, median, sdev*1e6/median, Quality))
    print(array)
    dmm.close()
    
F5700EP = rm.open_resource('GPIB0::1::INSTR')
    
F5700EP.write("OUT 250 V, 0 Hz")
time.sleep(5)
F5700EP.write("OUT 0 V, 0 Hz")
time.sleep(5)
F5700EP.write("STBY")    
ws['C373'] = float("%.2f" % ((time.time() - time_start)/60))
wb.save('test_4950.xlsx')
print("DCV time elapsed:%.2f minutes"
      % ((time.time() - time_start)/60))
time_start = time.time()
F5700EP.close()

        
########## OHM PERFORMANCE TEST ##########
print("OHM PERFORMANCE TEST")
F5700EP = rm.open_resource('GPIB0::1::INSTR')
MFC_OHM_list = ["1 OHM","1.9 OHM","10 OHM","19 OHM","100 OHM","190 OHM","1 KOHM","1.9 KOHM","10 KOHM",
                "19 KOHM","100 KOHM","190 KOHM","1 MOHM","1.9 MOHM","10 MOHM","19 MOHM","100 MOHM"]

F5700EP.write("OUT 0 OHM")
F5700EP.write("EXTSENSE ON")
F5700EP.write("OPER")
time.sleep(10)

F5700EP.close()
dmm = rm.open_resource('GPIB0::9::INSTR')
dmm.timeout = None
dmm.write("OHMS 10,PCENT_0,FWR,LCL_GUARD")
dmm.write("ZERO?")
time.sleep(100)

dmm.close()

for ix in range (0,17):
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None    
    array = []
    sdev = 0.0
    median = 0.0
    cutstr = []
    unc = 0
    Quality = 0
    res = MFC_OHM_list[ix]
    if ix == 0:
        dmm.write("OHMS 10,PCENT_10,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 10 OHM, PCENT_10")
    elif ix == 1:
        dmm.write("OHMS 10,PCENT_30,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 10 OHM, PCENT_30")
    elif ix == 2:
        dmm.write("OHMS 10,PCENT_100,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 10 OHM, PCENT_100")
    elif ix == 3:
        dmm.write("OHMS 10,PCENT_190,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 10 OHM, PCENT_190")
    elif ix == 4:
        dmm.write("OHMS 100,PCENT_100,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 100 OHM, PCENT_100")
    elif ix == 5:
        dmm.write("OHMS 100,PCENT_190,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 100 OHM, PCENT_190")
    elif ix == 6:
        dmm.write("OHMS 1000,PCENT_100,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 1 KOHM, PCENT_100")
    elif ix == 7:
        dmm.write("OHMS 1000,PCENT_190,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 1 KOHM, PCENT_190")
    elif ix == 8:
        dmm.write("OHMS 10000,PCENT_100,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 10 KOHM, PCENT_100")
    elif ix == 9:
        dmm.write("OHMS 10000,PCENT_190,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 10 KOHM, PCENT_190")
    elif ix == 10:
        dmm.write("OHMS 100000,PCENT_100,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 100 KOHM, PCENT_100")
    elif ix == 11:
        dmm.write("OHMS 100000,PCENT_190,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 100 KOHM, PCENT_190")
    elif ix == 12:
        dmm.write("OHMS 1000000,PCENT_100,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 1 MOHM, PCENT_100")
    elif ix == 13:
        dmm.write("OHMS 1000000,PCENT_190,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 1 MOHM, PCENT_190")
    elif ix == 14:
        dmm.write("OHMS 10000000,PCENT_100,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 10 MOHM, PCENT_100")
    elif ix == 15:
        dmm.write("OHMS 10000000,PCENT_190,FWR,LCL_GUARD")
        print("DMM OHMS FWR Range: 10 MOHM, PCENT_190")
    elif ix == 16:
        dmm.write("OHMS 100000000,PCENT_100,TWR,LCL_GUARD")
        dmm.close()
        F5700EP = rm.open_resource('GPIB0::1::INSTR')
        F5700EP.write("EXTSENSE OFF")
        F5700EP.close()
        dmm = rm.open_resource('GPIB0::9::INSTR')
        dmm.timeout = None    
        print("DMM OHMS TWR Range: 100 MOHM, PCENT_100") 
        
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
        
    F5700EP.write("OUT %s" % res) 
    F5700EP.write("OPER")
    
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None    
    if ix < 13 and ix >= 6:
        time.sleep(60)
    else:
        dmm.write("TRIG_SRCE INT")
        time.sleep(600) # waiting 600s at range 1M to 100M and 1 to 100R
        dmm.write("TRIG_SRCE EXT")
    for i in range (0,10):
        dmm.write("*TRG;GET;RDG?")
        res = float(dmm.read())
        array.extend([res])
    sdev = np.std(array[1:],ddof = 1)
    median = np.median(array[1:])
    
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    
    F5700EP.write("OUT?")
    res = F5700EP.read()
    cutstr = res.split(",")
    res = float(cutstr[0])
    ws['B' + str(105+ix)] = res
    ws['C' + str(105+ix)] = median
    ws['J' + str(105+ix)] = sdev*1e6/median
    F5700EP.write("UNCERT?")
    unc = F5700EP.read()
    cutstr = unc.split(",")
    ws['D' + str(105+ix)] = float(cutstr[0])
    
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None    
    dmm.write("DEVTN? READING")
    Quality = float(dmm.read())
    ws['K' + str(105+ix)] = Quality    
    print("Source = %s, dmm = %.8f ohm,  sdev = %.3f ppm, Quality(DMM) =  %s" % (res, median, sdev*1e6/median, Quality))
    print(array)
    dmm.close()
wb.save('test_4950.xlsx')

dmm = rm.open_resource('GPIB0::9::INSTR')
dmm.timeout = None    
dmm.write("OHMS 10,PCENT_0,FWR,LCL_GUARD")
### OHM zero 4w ###             

dmm.close()
F5700EP = rm.open_resource('GPIB0::1::INSTR')
                                                         
F5700EP.write("OUT 0 OHM")
F5700EP.write("EXTSENSE ON")
F5700EP.write("OPER")
F5700EP.close()
time.sleep(10)

for ix in range (0,8):
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None    
    array = []
    sdev = 0.0
    median = 0.0
    unc = 0
    Quality = 0
    if ix == 0:
        dmm.write("OHMS 10,PCENT_0,FWR,LCL_GUARD")
        print("DMM FWR Zero Range: 10 OHM")
    elif ix == 1:
        dmm.write("OHMS 100,PCENT_0,FWR,LCL_GUARD")
        print("DMM FWR Zero Range: 100 OHM")
    elif ix == 2:
        dmm.write("OHMS 1000,PCENT_0,FWR,LCL_GUARD")
        print("DMM FWR Zero Range: 1 KOHM")
    elif ix == 3:
        dmm.write("OHMS 10000,PCENT_0,FWR,LCL_GUARD")
        print("DMM FWR Zero Range: 10 KOHM")
    elif ix == 4:
        dmm.write("OHMS 100000,PCENT_0,FWR,LCL_GUARD")
        print("DMM FWR Zero Range: 100 KOHM")
    elif ix == 5:
        dmm.write("OHMS 1000000,PCENT_0,FWR,LCL_GUARD")
        print("DMM FWR Zero Range: 1 MOHM")
    elif ix == 6:
        dmm.write("OHMS 10000000,PCENT_0,FWR,LCL_GUARD")
        print("DMM FWR Zero Range: 10 MOHM")
    elif ix == 7:
        dmm.write("OHMS 100000000,PCENT_0,FWR,LCL_GUARD")
        print("DMM FWR Zero Range: 100 MOHM")
        
    for i in range (0,10):
        dmm.write("*TRG;GET;RDG?")
        res = float(dmm.read())
        array.extend([res])
    sdev = np.std(array[1:],ddof = 1)
    median = np.median(array[1:])
    ws['C' + str(130+ix)] = median
    ws['J' + str(130+ix)] = sdev
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    F5700EP.write("UNCERT?")
    unc = F5700EP.read()
    
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    
    cutstr = unc.split(",")
    ws['D' + str(130+ix)] = float(cutstr[0])
    dmm.write("DEVTN? ABSOLUTE")
    Quality = float(dmm.read())
    ws['K' + str(130+ix)] = Quality    
    print("Source = %s, dmm = %.8f ohm,  sdev = %.8f ohm,  Quality(DMM) = %.8f ohm " % ("0 ohm", median, sdev, Quality))
    print(array)
    dmm.close()
wb.save('test_4950.xlsx')


dmm = rm.open_resource('GPIB0::9::INSTR')
dmm.timeout = None 
dmm.write("OHMS 10,PCENT_0,TWR,LCL_GUARD")
dmm.close()
F5700EP = rm.open_resource('GPIB0::1::INSTR')
### OHM zero 2w ###
F5700EP.write("OUT 0 OHM")
F5700EP.write("EXTSENSE OFF")
F5700EP.write("OPER")
F5700EP.close()
time.sleep(10)
for ix in range (0,8):
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    array = []
    sdev = 0.0
    median = 0.0
    unc = 0
    Quality = 0

    if ix == 0:
        dmm.write("OHMS 10,PCENT_0,TWR,LCL_GUARD")
        print("DMM OHMS TWR Zero Range: 10 OHM")
    elif ix == 1:
        dmm.write("OHMS 100,PCENT_0,TWR,LCL_GUARD")
        print("DMM OHMS TWR Zero Range: 100 OHM")
    elif ix == 2:
        dmm.write("OHMS 1000,PCENT_0,TWR,LCL_GUARD")
        print("DMM OHMS TWR Zero Range: 1 KOHM")
    elif ix == 3:
        dmm.write("OHMS 10000,PCENT_0,TWR,LCL_GUARD")
        print("DMM OHMS TWR Zero Range: 10 KOHM")
    elif ix == 4:
        dmm.write("OHMS 100000,PCENT_0,TWR,LCL_GUARD")
        print("DMM OHMS TWR Zero Range: 100 KOHM")
    elif ix == 5:
        dmm.write("OHMS 1000000,PCENT_0,TWR,LCL_GUARD")
        print("DMM OHMS TWR Zero Range: 1 MOHM")
    elif ix == 6:
        dmm.write("OHMS 10000000,PCENT_0,TWR,LCL_GUARD")
        print("DMM OHMS TWR Zero Range: 10 MOHM")        
    elif ix == 7:
        dmm.write("OHMS 100000000,PCENT_0,TWR,LCL_GUARD")
        print("DMM OHMS TWR Zero Range: 100 MOHM")
        
    for i in range (0,10):
        dmm.write("*TRG;GET;RDG?")
        res = float(dmm.read())
        array.extend([res])
    sdev = np.std(array[1:],ddof = 1)
    median = np.median(array[1:])

    ws['C' + str(146+ix)] = median
    ws['J' + str(146+ix)] = sdev
    
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    
    F5700EP.write("UNCERT?")
    unc = F5700EP.read()
    cutstr = unc.split(",")
    ws['D' + str(146+ix)] = float(cutstr[0])
    
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    
    dmm.write("DEVTN? ABSOLUTE")
    Quality = float(dmm.read())
    ws['K' + str(146+ix)] = Quality    
    print("Source = %s, dmm = %.8f ohm,  sdev = %.8f ohm, Quality(DMM) = %.8f ohm " % ("0 ohm", median, sdev, Quality))
    print(array)
    dmm.close()

F5700EP = rm.open_resource('GPIB0::1::INSTR')
F5700EP.write("EXTSENSE OFF")    
ws['D373'] = float("%.2f" % ((time.time() - time_start)/60))
wb.save('test_4950.xlsx')
print("OHM time elapsed:%.2f minutes"
      % ((time.time() - time_start)/60))

time_start = time.time()
F5700EP.write("OUT 0 V, 0 Hz")
F5700EP.write("STBY")

F5700EP.close()


#########################
#ACV FWR
while 1:
    str_input = input("connect the cable for ACV FWR, and then input: go\n")
    if (str_input == 'go'):
        break
    else:
        print("input again")
#########################
   
########## ACV PERFORMANCE TEST ##########
print("ACV PERFORMANCE TEST")
ACV_LIN_list = ["1 V,1 KHz","2 V, 1 KHz","5 V, 1 KHz","10 V, 1 KHz","12 V, 1 KHz","15 V, 1 KHz","19 V, 1 KHz"]
### AC VOLTAGE Linearity Checks ###

dmm = rm.open_resource('GPIB0::9::INSTR')
dmm.timeout = None 
dmm.write("ACV 10,FREQ_1k,PCENT_100,FWR,LCL_GUARD")
dmm.close()
F5700EP = rm.open_resource('GPIB0::1::INSTR')

F5700EP.write("OUT 1 V, 1 KHz")
F5700EP.write("OPER")
time.sleep(180)# thermal sensor warm up
F5700EP.close()

for ix in range (0,7):
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    array = []
    sdev = 0.0
    median = 0.0
    unc = 0
    Quality = 0
    acv = ACV_LIN_list[ix]        
    F5700EP.write("OUT %s" % acv)
    F5700EP.write("OPER")
    
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    
    time.sleep(60)
    for i in range (0,10):
        dmm.write("*TRG;GET;RDG?")
        volt = float(dmm.read())
        array.extend([volt])
    sdev = np.std(array[1:],ddof = 1)
    median = np.median(array[1:])
    ws['C' + str(161+ix)] = median
    ws['J' + str(161+ix)] = sdev*1e6/median
    
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    
    F5700EP.write("UNCERT?")
    unc = F5700EP.read()
    cutstr = unc.split(",")
    if str(cutstr[1]) == "PCT":
        ws['D' + str(161+ix)] = float(cutstr[0])*1e4
    else:
        ws['D' + str(161+ix)] = float(cutstr[0])
        
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
        
    dmm.write("DEVTN? READING")
    Quality = float(dmm.read())
    ws['K' + str(161+ix)] = Quality    
    print("Source = %s , dmm = %.8f Vac,  sdev = %.3f ppm, Quality(DMM) =  %s" % (acv, median, sdev*1e6/median, Quality))
    print(array)
    dmm.close()
    
F5700EP = rm.open_resource('GPIB0::1::INSTR')
F5700EP.write("STBY")
wb.save('test_4950.xlsx')

### ACV SYNC Mode ###                                                                   
ACV_list = ["0.001 V, 10 Hz","0.001 V, 20 Hz","0.001 V, 30 Hz","0.001 V, 40 Hz","0.001 V, 55 Hz","0.001 V, 300 Hz","0.001 V, 1 KHz","0.001 V, 10 KHz","0.001 V, 20 KHz","0.001 V, 30 KHz","0.001 V, 50 KHz",
            "0.001 V, 100 KHz","0.001 V, 200 KHz","0.001 V, 300 KHz","0.001 V, 500 KHz","0.001 V, 1 MHz",
            "0.01 V, 10 Hz","0.01 V, 20 Hz","0.01 V, 30 Hz","0.01 V, 40 Hz","0.01 V, 55 Hz","0.01 V, 300 Hz","0.01 V, 1 KHz","0.01 V, 10 KHz","0.01 V, 20 KHz","0.01 V, 30 KHz","0.01 V, 50 KHz",
            "0.01 V, 100 KHz","0.01 V, 200 KHz","0.01 V, 300 KHz","0.01 V, 500 KHz","0.01 V, 1 MHz",
            "0.1 V, 10 Hz","0.1 V, 20 Hz","0.1 V, 30 Hz","0.1 V, 40 Hz","0.1 V, 55 Hz","0.1 V, 300 Hz","0.1 V, 1 KHz","0.1 V, 10 KHz","0.1 V, 20 KHz","0.1 V, 30 KHz","0.1 V, 50 KHz",
            "0.1 V, 100 KHz","0.1 V, 200 KHz","0.1 V, 300 KHz","0.1 V, 500 KHz","0.1 V, 1 MHz",
            "1 V, 10 Hz","1 V, 20 Hz","1 V, 30 Hz","1 V, 40 Hz","1 V, 55 Hz","1 V, 300 Hz","1 V, 1 KHz","1 V, 10 KHz","1 V, 20 KHz","1 V, 30 KHz","1 V, 50 KHz",
            "1 V, 100 KHz","1 V, 200 KHz","1 V, 300 KHz","1 V, 500 KHz","1 V, 1 MHz",           
            "10 V, 10 Hz","10 V, 20 Hz","10 V, 30 Hz","10 V, 40 Hz","10 V, 55 Hz","10 V, 300 Hz","10 V, 1 KHz","10 V, 10 KHz","10 V, 20 KHz","10 V, 30 KHz","10 V, 50 KHz",
            "10 V, 100 KHz","10 V, 200 KHz","10 V, 300 KHz","10 V, 500 KHz","10 V, 1 MHz",
            "19 V, 10 Hz","19 V, 20 Hz","19 V, 30 Hz","19 V, 40 Hz","19 V, 55 Hz","19 V, 300 Hz","19 V, 1 KHz","19 V, 10 KHz","19 V, 20 KHz","19 V, 30 KHz","19 V, 50 KHz",
            "19 V, 100 KHz","19 V, 200 KHz","19 V, 300 KHz","19 V, 500 KHz","19 V, 1 MHz", 
            "100 V, 10 Hz","100 V, 20 Hz","100 V, 30 Hz","100 V, 40 Hz","100 V, 55 Hz","100 V, 300 Hz","100 V, 1 KHz","100 V, 10 KHz","100 V, 20 KHz","100 V, 30 KHz","100 V, 50 KHz",
            "100 V, 100 KHz","100 V, 200 KHz","1000 V, 55 Hz","1000 V, 300 Hz","700 V, 1 KHz"]


F5700EP.close()
dmm = rm.open_resource('GPIB0::9::INSTR')
dmm.timeout = None 

dmm.write("ACV 10,FREQ_1k,PCENT_100,FWR,LCL_GUARD")
for ix in range (0,112):
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    array = []
    sdev = 0.0
    median = 0.0
    cutstr = []
    unc = 0
       
    acv = ACV_list[ix]
    cutstr = acv.split(" ")
    if cutstr[0] == "19":
        pcent = "PCENT_190"
    else:
        pcent = "PCENT_100"
    if cutstr[3] == "Hz":
        dmm.write("ACV %s,FREQ_%s,%s,FWR,LCL_GUARD" %(cutstr[0],cutstr[2],pcent))
    elif cutstr[3] == "KHz":
        dmm.write("ACV %s,FREQ_%sk,%s,FWR,LCL_GUARD" %(cutstr[0],cutstr[2],pcent))
    elif cutstr[3] == "MHz":
        dmm.write("ACV %s,FREQ_%sM,%s,FWR,LCL_GUARD" %(cutstr[0],cutstr[2],pcent))
    else:
        dmm.write("ACV 1000,FREQ_1k,PCENT_100,FWR,LCL_GUARD")
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    
    F5700EP.write("FAULT?")
    fault = F5700EP.read()
    fault = ''.join(e for e in fault if e.isalnum())
    if fault != "0":
        print("Calibrator fault "+fault+" it will probably refuse to operate now.")
        #F5700EP.write("EXPLAIN?")
        #print(F5700EP.read())
    
    F5700EP.write("OUT %s" % acv)
    F5700EP.write("OPER")
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    if ix == 109:
        dmm.write("TRIG_SRCE INT")
        time.sleep(600)#wait 10 minitue for the first 1kV testing point.
        dmm.write("TRIG_SRCE EXT")
    else:
        time.sleep(10) 
    for i in range (0,10):
        dmm.write("*TRG;GET;RDG?")
        volt = float(dmm.read())
        array.extend([volt])
    sdev = np.std(array[1:],ddof = 1)
    median = np.median(array[1:])
    ws['C' + str(170+ix)] = median
    ws['J' + str(170+ix)] = sdev*1e6/median
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    F5700EP.write("UNCERT?")
    unc = F5700EP.read()
    cutstr = unc.split(",")
    if cutstr[1] == "PCT":
        ws['D' + str(170+ix)] = float(cutstr[0])*1e4
    else:
        ws['D' + str(170+ix)] = float(cutstr[0])
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    dmm.write("DEVTN? READING")
    Quality = float(dmm.read())
    ws['K' + str(170+ix)] = Quality        
    print("Source = %s , dmm = %.8f Vac,  sdev = %.3f ppm, Quality(DMM) =  %s" % (acv, median, sdev*1e6/median, Quality))
    print(array)
    dmm.close()
    
    
F5700EP = rm.open_resource('GPIB0::1::INSTR') 
F5700EP.write("OUT 250 V, 1 KHz")
time.sleep(5)
F5700EP.write("OUT 0 V, 0 Hz")
time.sleep(5)
F5700EP.write("STBY")

ws['E373'] = float("%.2f" % ((time.time() - time_start)/60))
wb.save('test_4950.xlsx')
print("ACV time elapsed:%.2f minutes"
      %((time.time() - time_start)/60))


#########################
#remove all cable and short bar, connect a Twisted pair cable for current testing
while 1:
    str_input = input("Reconnect the cable for Current testing, and then input: go\n")
    if (str_input == 'go'):
        break
    else:
        print("input again")
#########################

   
time_start = time.time()        
########## DCI test 100uA to 1A ##########
print("DCI PERFORMANCE TEST")
DCI_list = ["1 uA","10 uA","50 uA","100 uA","-100 uA","-50 uA","-10 uA",
            "0.1 mA","0.5 mA","1 mA","-1 mA","-0.5 mA",
            "1 mA","5 mA","10 mA","-10 mA","-5 mA",
            "10 mA","50 mA","100 mA","-100 mA","-50 mA",
            "0.1 A","0.5 A","-0.5 A","-1 A","1 A"]

F5700EP.close()
dmm = rm.open_resource('GPIB0::9::INSTR')
dmm.timeout = None 
dmm.write("DCI 1,PCENT_100,LCL_GUARD")
dmm.close()
F5700EP = rm.open_resource('GPIB0::1::INSTR')
F5700EP.write("OUT 0 uA, 0 Hz")
F5700EP.write("CUR_POST AUX")
F5700EP.close()
for ix in range (0,27):
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    array = []
    sdev = 0.0
    median = 0.0
    unc = 0

    iout = DCI_list[ix]
    if ix < 7:
        dmm.write("DCI 0.0001,PCENT_100,LCL_GUARD")
        print("DMM DCI Range: 100uA")
    elif ix >= 7 and ix < 12:
        dmm.write("DCI 0.001,PCENT_100,LCL_GUARD")
        print("DMM DCI Range: 1mA")
    elif ix >= 12 and ix < 17:
        dmm.write("DCI 0.01,PCENT_100,LCL_GUARD")
        print("DMM DCI Range: 10mA")
    elif ix >= 17 and ix < 22:
        dmm.write("DCI 0.1,PCENT_100,LCL_GUARD")
        print("DMM DCI Range: 100mA")
    elif ix >= 22 and ix < 27:
        dmm.write("DCI 1,PCENT_100,LCL_GUARD")
        print("DMM DCI Range: 1A")        
    else:
        dmm.write("DCI 1,PCENT_100,LCL_GUARD")
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    F5700EP.write("OUT %s" % iout)
    F5700EP.write("OPER")
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    if ix == 23:
        time.sleep(300)
    elif ix == 25:
        time.sleep(600)
    else:
        time.sleep(10)
    for i in range (0,5):
        dmm.write("*TRG;GET;RDG?")
        volt = float(dmm.read())
        array.extend([volt])
    sdev = np.std(array[1:],ddof = 1)
    median = np.median(array[1:])
    ws['C' + str(287+ix)] = median
    ws['J' + str(287+ix)] = sdev*1e6/median
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    F5700EP.write("UNCERT?")
    unc = F5700EP.read()
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    cutstr = unc.split(",")
    if str(cutstr[1]) == "A":
        ws['D' + str(287+ix)] = float(cutstr[0])*1e6/median
    else:    
        ws['D' + str(287+ix)] = float(cutstr[0])
    dmm.write("DEVTN? READING")
    Quality = float(dmm.read())
    ws['K' + str(287+ix)] = Quality
    print("Source = %s, dmm = %.10f A,  sdev = %.3f ppm, Quality(DMM) = %s" % (iout, median, sdev*1e6/median, Quality))
    print(array)
    dmm.close()
wb.save('test_4950.xlsx')

F5700EP = rm.open_resource('GPIB0::1::INSTR')
F5700EP.write("OUT 0 uA, 0 Hz")
F5700EP.close()
dmm = rm.open_resource('GPIB0::9::INSTR')
dmm.timeout = None 

ws['F373'] = float("%.2f" % ((time.time() - time_start)/60))
wb.save('test_4950.xlsx')
print("DCI time elapsed:%.2f minutes"
      % ((time.time() - time_start)/60))
time_start = time.time()

########## ACI test 10uA to 1A ##########
ACI_list = ["10 uA, 10 Hz","10 uA, 20 Hz","10 uA, 30 Hz","10 uA, 40 Hz","10 uA, 55 Hz","10 uA, 300 Hz","10 uA, 1 KHz","10 uA, 5 KHz","10 uA, 10 KHz",
            "100 uA, 10 Hz","100 uA, 20 Hz","100 uA, 30 Hz","100 uA, 40 Hz","100 uA, 55 Hz","100 uA, 300 Hz","100 uA, 1 KHz","100 uA, 5 KHz","100 uA, 10 KHz",
            "1 mA, 10 Hz","1 mA, 20 Hz","1 mA, 30 Hz","1 mA, 40 Hz","1 mA, 55 Hz","1 mA, 300 Hz","1 mA, 1 KHz","1 mA, 5 KHz","1 mA, 10 KHz",
            "10 mA, 10 Hz","10 mA, 20 Hz","10 mA, 30 Hz","10 mA, 40 Hz","10 mA, 55 Hz","10 mA, 300 Hz","10 mA, 1 KHz","10 mA, 5 KHz","10 mA, 10 KHz",
            "100 mA, 10 Hz","100 mA, 20 Hz","100 mA, 30 Hz","100 mA, 40 Hz","100 mA, 55 Hz","100 mA, 300 Hz","100 mA, 1 KHz","100 mA, 5 KHz","100 mA, 10 KHz",
            "1 A, 10 Hz","1 A, 20 Hz","1 A, 30 Hz","1 A, 40 Hz","1 A, 55 Hz","1 A, 300 Hz","1 A, 1 KHz","1 A, 5 KHz","1 A, 10 KHz"]

dmm.write("ACI 1,FREQ_1k,LCL_GUARD")
dmm.close()
for ix in range (0,54):
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    array = []
    sdev = 0.0
    median = 0.0
    unc = 0
    ACIrange = 0
    
    iout = ACI_list[ix]
    cutstr = iout.split(" ")
    if cutstr[1] == "uA,":
        if float(cutstr[0]) <= 199:
            ACIrange = "0.0001"
    elif cutstr[1] == "mA,":
        if float(cutstr[0]) <= 1.999:
            ACIrange = "0.001"
        elif (float(cutstr[0]) >= 2) and (float(cutstr[0]) <=19.999):
            ACIrange = "0.01"
        elif (float(cutstr[0]) >= 20) and (float(cutstr[0]) < 199.999):
            ACIrange = "0.1"
        else:
            ACIrange = "1"
    elif cutstr[1] == "A,":
        ACIrange = "1"
    else:
        ACIrange = "1"
  
    if cutstr[3] == "Hz":
        dmm.write("ACI %s,FREQ_%s,LCL_GUARD" %(ACIrange,cutstr[2]))
    elif cutstr[3] == "KHz":
        dmm.write("ACI %s,FREQ_%sk,LCL_GUARD" %(ACIrange,cutstr[2]))
    else:
        dmm.write("ACI 1,FREQ_1k,LCL_GUARD")        
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    F5700EP.write("OUT %s" % iout)
    F5700EP.write("OPER")
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    if ix == 45:
        dmm.write("TRIG_SRCE INT")
        time.sleep(600)
        dmm.write("TRIG_SRCE EXT")        
    else:
        time.sleep(10)
    for i in range (0,5):
        dmm.write("*TRG;GET;RDG?")
        volt = float(dmm.read())
        array.extend([volt])
    sdev = np.std(array[1:],ddof = 1)
    median = np.median(array[1:])
    ws['C' + str(317+ix)] = median
    ws['J' + str(317+ix)] = sdev*1e6/median
    dmm.close()
    F5700EP = rm.open_resource('GPIB0::1::INSTR')
    F5700EP.write("UNCERT?")
    unc = F5700EP.read()
    F5700EP.close()
    dmm = rm.open_resource('GPIB0::9::INSTR')
    dmm.timeout = None 
    cutstr = unc.split(",")
    if str(cutstr[1]) == "A":
        ws['D' + str(317+ix)] = float(cutstr[0])*1e6/median
    else:
        ws['D' + str(317+ix)] = float(cutstr[0])
    dmm.write("DEVTN? READING")
    Quality = float(dmm.read())
    ws['K' + str(317+ix)] = Quality
    print("Source = %s, dmm = %.10f A,  sdev = %.3f ppm, Quality(DMM) =  %s" % (iout, median, sdev*1e6/median, Quality)) 
    print(array)
    dmm.close()

ws['G373'] = float("%.2f" % ((time.time() - time_start)/60))
wb.save('test_4950.xlsx')
print("ACI time elapsed:%.2f minutes"
      % ((time.time() - time_start)/60))

#Calibration Test Results Summary

wb.save('test_4950.xlsx')

#Reset the DMM and MFC####
F5700EP = rm.open_resource('GPIB0::1::INSTR')
F5700EP.write("OUT 0 V, 0 Hz")
F5700EP.write("STBY")
F5700EP.write("*RST")
F5700EP.write("*CLS")
F5700EP.write("OUT 0 V, 0 Hz")     
F5700EP.write("RANGELCK ON")
F5700EP.close()
dmm = rm.open_resource('GPIB0::9::INSTR')
dmm.write("*RST")
dmm.close()





